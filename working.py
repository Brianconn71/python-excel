from openpyxl import Workbook, load_workbook

wb = load_workbook('GTN.xlsx')
ws = wb.active
print(ws['A2'].value)
ws['A2'].value = "Test"
print(ws['A2'].value)

wb.save('GTN.xlsx')